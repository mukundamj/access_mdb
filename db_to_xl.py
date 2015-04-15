import pypyodbc
import datetime
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter


pypyodbc.lowercase = False
conn = pypyodbc.connect(
    r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};" +
    r"Dbq=C:\DB_PY\Lysimeter.accdb;")
cur = conn.cursor()

five_min = datetime.timedelta(0,300)

#For SE Lysimeter

start =datetime.datetime(2015,03,28,00,00)
a=start
b=start+datetime.timedelta(0,240)
wb = load_workbook('SE_Lysimeter_TDT_Sensors_Experiment-02_03_28_2015.xlsx')
ws = wb.active
  

r1=r2=r3=r4=r5=r6=r7=r8=r9=r10=r11=r12=r13=r14=r15=r16=7
r17=r18=r19=r20=r21=r22=r23=r24=r25=r26=r27=r28=r29=r30=r31=r32=7

while a < (start+datetime.timedelta(1)):
    params = (a,b)
    sql = "SELECT Date_Time,Address,Volumetric_Water_Content, Soil_Temperature FROM SE_Data_5 WHERE Date_Time >= ? AND Date_Time <= ?"
    for row in cur.execute(sql,params):
        flag = 1
        if(str(row.get('Address')) == 'A'):
            c = 4
            r = r1=r1+1
        elif(str(row.get('Address')) == 'B'):
            c = 31
            r = r2=r2+1
        elif(str(row.get('Address')) == 'C'):
            c = 22
            r = r3=r3+1
        elif(str(row.get('Address')) == 'D'):
            c = 13
            r = r4=r4+1
        elif(str(row.get('Address')) == 'E'):
            c = 53
            r = r5=r5+1
        elif(str(row.get('Address')) == 'F'):
            c = 2
            r = r6=r6+1
        elif(str(row.get('Address')) == 'G'):
            c = 17
            r = r7=r7+1
        elif(str(row.get('Address')) == 'H'):
            c = 20
            r = r8=r8+1
        elif(str(row.get('Address')) == 'I'):
            c = 67
            r = r9=r9+1
        elif(str(row.get('Address')) == 'J'):
            c = 58
            r = r10=r10+1
        elif(str(row.get('Address')) == 'K'):
            c = 56
            r = r11=r11+1
        elif(str(row.get('Address')) == 'L'):
            c = 65
            r = r12=r12+1
        elif(str(row.get('Address')) == 'M'):
            c = 35
            r = r13=r13+1
        elif(str(row.get('Address')) == 'N'):
            c = 29
            r = r14=r14+1
        elif(str(row.get('Address')) == 'O'):
            c = 8
            r = r15=r15+1
        elif(str(row.get('Address')) == 'P'):
            c = 6
            r = r16=r16+1
        elif(str(row.get('Address')) == 'Q'):
            c = 38
            r = r17=r17+1
        elif(str(row.get('Address')) == 'R'):
            c = 62
            r = r18=r18+1
        elif(str(row.get('Address')) == 'S'):
            c = 49
            r = r19=r19+1
        elif(str(row.get('Address')) == 'T'):
            c = 11
            r = r20=r20+1
        elif(str(row.get('Address')) == 'U'):
            c = 44
            r = r21=r21+1
        elif(str(row.get('Address')) == 'V'):
            c = 71
            r = r22=r22+1
        elif(str(row.get('Address')) == 'W'):
            c = 33
            r = r23=r23+1
        elif(str(row.get('Address')) == 'X'):
            c = 26
            r = r24=r24+1
        elif(str(row.get('Address')) == 'Y'):
            c = 47
            r = r25=r25+1
        elif(str(row.get('Address')) == 'Z'):
            c = 51
            r = r26=r26+1
        elif(str(row.get('Address')) == 'a'):
            c = 42
            r = r27=r27+1
        elif(str(row.get('Address')) == 'b'):
            c = 60
            r = r28=r28+1
        elif(str(row.get('Address')) == 'c'):
            c = 24
            r = r29=r29+1
        elif(str(row.get('Address')) == 'd'):
            c = 40
            r = r30=r30+1
        elif(str(row.get('Address')) == 'e'):
            c = 15
            r = r31=r31+1
        elif(str(row.get('Address')) == 'f'):
            c = 69
            r = r32=r32+1
        else:
            continue
        ws.cell(row = r,column=1).value = row.get('Date_Time')
        ws.cell(row = r,column=c).value = row.get('Volumetric_Water_Content')
        ws.cell(row = r,column=c+1).value = row.get('Soil_Temperature')
#       print(u"Date_Time is {0} Address is {1} moisture is {2} temp is {3}".format(row.get('Date_Time'),row.get('Address'),row.get('Volumetric_Water_Content'), row.get('Soil_Temperature')))
    a = a+five_min
    b = b+five_min   
wb.save('SE_Lysimeter_TDT_Sensors_Experiment-02_03_28_2015.xlsx')


#For SW Lysimeter

start = datetime.datetime(2015,03,28,00,00)
a=start
b=start+datetime.timedelta(0,240)
wb = load_workbook('SW_Lysimeter_TDT_Sensors_Experiment-02_03_28_2015_MT.xlsx')
ws = wb.active

r1=r2=r3=r4=r5=r6=r7=r8=r9=r10=r11=r12=r13=r14=r15=r16=7
r17=r18=r19=r20=r21=r22=r23=r24=r25=r26=r27=r28=r29=r30=r31=r32=7

while a < (start+datetime.timedelta(1)):
    params = (a,b)
    sql = "SELECT Date_Time,Address,Volumetric_Water_Content, Soil_Temperature FROM SW_Data_5 WHERE Date_Time >= ? AND Date_Time <= ?"
    for row in cur.execute(sql,params):
        flag = 1
        if(str(row.get('Address')) == 'A'):
            c = 20
            r = r1=r1+1
        elif(str(row.get('Address')) == 'B'):
            c = 47
            r = r2=r2+1
        elif(str(row.get('Address')) == 'C'):
            c = 2
            r = r3=r3+1
        elif(str(row.get('Address')) == 'D'):
            c = 11
            r = r4=r4+1
        elif(str(row.get('Address')) == 'E'):
            c = 33
            r = r5=r5+1
        elif(str(row.get('Address')) == 'F'):
            c = 6
            r = r6=r6+1
        elif(str(row.get('Address')) == 'G'):
            c = 38
            r = r7=r7+1
        elif(str(row.get('Address')) == 'H'):
            c = 56
            r = r8=r8+1
        elif(str(row.get('Address')) == 'I'):
            c = 40
            r = r9=r9+1
        elif(str(row.get('Address')) == 'J'):
            c = 58
            r = r10=r10+1
        elif(str(row.get('Address')) == 'K'):
            c = 24
            r = r11=r11+1
        elif(str(row.get('Address')) == 'L'):
            c = 15
            r = r12=r12+1
        elif(str(row.get('Address')) == 'M'):
            c = 29
            r = r13=r13+1
        elif(str(row.get('Address')) == 'N'):
            c = 31
            r = r14=r14+1
        elif(str(row.get('Address')) == 'O'):
            c = 65
            r = r15=r15+1
        elif(str(row.get('Address')) == 'P'):
            c = 4
            r = r16=r16+1
        elif(str(row.get('Address')) == 'Q'):
            c = 69
            r = r17=r17+1
        elif(str(row.get('Address')) == 'R'):
            c = 67
            r = r18=r18+1
        elif(str(row.get('Address')) == 'S'):
            c = 22
            r = r19=r19+1
        elif(str(row.get('Address')) == 'T'):
            c = 60
            r = r20=r20+1
        elif(str(row.get('Address')) == 'U'):
            c = 13
            r = r21=r21+1
        elif(str(row.get('Address')) == 'V'):
            c = 51
            r = r22=r22+1
        elif(str(row.get('Address')) == 'W'):
            c = 35
            r = r23=r23+1
        elif(str(row.get('Address')) == 'X'):
            c = 44
            r = r24=r24+1
        elif(str(row.get('Address')) == 'Y'):
            c = 26
            r = r25=r25+1
        elif(str(row.get('Address')) == 'Z'):
            c = 49
            r = r26=r26+1
        elif(str(row.get('Address')) == 'a'):
            c = 62
            r = r27=r27+1
        elif(str(row.get('Address')) == 'b'):
            c = 42
            r = r28=r28+1
        elif(str(row.get('Address')) == 'c'):
            c = 8
            r = r29=r29+1
        elif(str(row.get('Address')) == 'd'):
            c = 53
            r = r30=r30+1
        elif(str(row.get('Address')) == 'e'):
            c = 17
            r = r31=r31+1
        elif(str(row.get('Address')) == 'f'):
            c = 71
            r = r32=r32+1
        else:
            continue
        ws.cell(row = r,column=1).value = row.get('Date_Time')
        ws.cell(row = r,column=c).value = row.get('Volumetric_Water_Content')
        ws.cell(row = r,column=c+1).value = row.get('Soil_Temperature')
#       print(u"Date_Time is {0} Address is {1} moisture is {2} temp is {3}".format(row.get('Date_Time'),row.get('Address'),row.get('Volumetric_Water_Content'), row.get('Soil_Temperature')))
    a = a+five_min
    b = b+five_min

wb.save('SW_Lysimeter_TDT_Sensors_Experiment-02_03_28_2015_MT.xlsx')
cur.close()
conn.close()

#"ABCDEFGHIJKLMNOPQRSTUVWXYXabcde":
#ctrl+[ and ctrl+] for left and right indent
#alt+3 for inserting block comment
#alt+4 for inserting block comment

